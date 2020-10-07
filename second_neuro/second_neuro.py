import numpy as np
import openpyxl
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.styles import PatternFill

""" CONST """
TICKRATE = 0.01
PATH = "F://Projects//neuro//"

def read_file(N = 0):
    first_signal_opened = []
    second_signal_opened = []
    third_signal_opened = []
    values_list = None
    with open(PATH + 'openedEyes.asc') as file:
        for elem in file.read().split('\n'):
            values_list = elem.split(' ')
            if len(values_list) == 3:
                first_signal_opened.append(int(values_list[0]))
                second_signal_opened.append(int(values_list[1]))
                third_signal_opened.append(int(values_list[2]))
                if len(first_signal_opened) == N and N != 0: 
                    break
    first_signal_closed = []
    second_signal_closed = []
    third_signal_closed= []
    values_list = None
    with open(PATH + 'closedEyes.asc') as file:
        for elem in file.read().split('\n'):
            values_list = elem.split(' ')
            if len(values_list) == 3:
                first_signal_closed.append(int(values_list[0]))
                second_signal_closed.append(int(values_list[1]))
                third_signal_closed.append(int(values_list[2]))
                if len(first_signal_closed) == N and N != 0: 
                    break
    signal_list = [first_signal_opened, second_signal_opened, third_signal_opened, \
    first_signal_closed, second_signal_closed, third_signal_closed]
    return signal_list

def correlation(first_signal, second_signal):
    first_signal_average = 0
    second_signal_average = 0
    for i in range(0, len(first_signal)):
        first_signal_average += first_signal[i]
    first_signal_average /= len(first_signal)
    for i in range(0, len(second_signal)):
        second_signal_average += second_signal[i]
    second_signal_average /= len(second_signal)
    numerator = 0
    denominator_left = 0
    denominator_right = 0
    for i in range(0, len(first_signal)):
        numerator += (first_signal[i] - first_signal_average) * (second_signal[i] - second_signal_average)
        denominator_left += ((first_signal[i] - first_signal_average) ** 2)
        denominator_right += ((second_signal[i] - second_signal_average) ** 2)
    return numerator/np.sqrt(denominator_left * denominator_right)

if __name__ == "__main__":
    wb = openpyxl.load_workbook(filename = 'F:\\Projects\\neuro\\second_neuro\\correlation.xlsx')
    signal_name_list = ['Opened First Signal', 'Opened Second Signal', 'Opened Third Signal', 'Closed First Signal', 'Closed Second Signal', 'Closed Third Signal']
    sheet = wb['Лист1']
    signal_list = read_file(2000)
    time = [0.1, 0.2, 0.5, 1, 2, 3]
    row_count = 0
    item_count = 0
    step = 0
    ticks = []
    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='FF0000'),
        right=Side(border_style=BORDER_THIN, color='FF0000'),
        top=Side(border_style=BORDER_THIN, color='FF0000'),
        bottom=Side(border_style=BORDER_THIN, color='FF0000')
    )
    for item in time:
        ticks.append(item/TICKRATE)
    for item in ticks:
        for k in range(0, len(signal_list[0]), int(item)):
            for u in range(1, 8):
                if (u == 1): 
                    sheet.cell(row = 1 + row_count , column= 1 + item_count).value = f'{np.round(item/100 * step, 1)} s'
                    sheet.cell(row = 1 + row_count , column= 1 + item_count).border = thin_border
                    sheet.cell(row = 1 + row_count , column= 1 + item_count).fill = \
                        PatternFill(start_color='FFEE08', end_color='FFEE08', fill_type = 'solid')
                else:
                    sheet.cell(row = u + row_count, column= 1 + item_count).value = signal_name_list[u-2]
                    sheet.cell(row = u + row_count, column= 1 + item_count).border = thin_border
                    sheet.cell(row = u + row_count, column= 1 + item_count).fill = \
                        PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type = 'solid')
            for m in range(2, 8):
                sheet.cell(row = 1 + row_count, column= m + item_count).value = signal_name_list[m-2]
                sheet.cell(row = 1 + row_count, column= m + item_count).border = thin_border
                sheet.cell(row = 1 + row_count, column= m + item_count).fill = \
                    PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type = 'solid')
            for i in range(2, 8):
                for j in range(2, 8):
                    if i == j:
                        sheet.cell(row = i + row_count, column= j + item_count).value = '----------------------------'
                    else:
                        if len(signal_list[i-2][int(item * k):int(item * (k + 1))]) != 0:
                            sheet.cell(row = i + row_count, column = j + item_count).value = \
                            correlation(signal_list[i-2][int(item * k):int(item * (k + 1))], signal_list[j-2][int(item * k):int(item * (k + 1))])
            item_count += 7
            step += 1
        row_count += 7
        item_count = 0
        step = 0
    wb.save('F:\\Projects\\neuro\\second_neuro\\correlation.xlsx')